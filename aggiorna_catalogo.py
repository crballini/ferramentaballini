import pandas as pd
import json
import os
import re
import argparse
import copy
import shutil
import openpyxl
import urllib.request
import urllib.parse
import mimetypes
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Mappatura dei Reparti del sito ai file JSON corrispondenti in assets/data/
REPARTO_MAP = {
    'Edilizia e ferramenta': 'prodotti-edilizia-ferramenta.json',
    'Elettricità': 'prodotti-elettricita.json',
    'Elettrodomestici': 'prodotti-elettrodomestici.json',
    'Elettroutensili': 'prodotti-elettroutensili.json',
    'Idraulica': 'prodotti-idraulica.json',
    'Utensili manuali': 'prodotti-utensili-manuali.json',
    'Vernici e colori': 'prodotti-vernici-colori.json',
    'Vario': 'prodotti-vario.json'
}

# Prefissi dei Reparti per i Codici Prodotto (2 lettere)
REPARTO_PREFIX = {
    'Edilizia e ferramenta': 'EF',
    'Elettricità': 'EL',
    'Elettrodomestici': 'ED',
    'Elettroutensili': 'EU',
    'Idraulica': 'ID',
    'Utensili manuali': 'UM',
    'Vernici e colori': 'VC',
    'Vario': 'VA'
}

# Cartelle "Web-Safe" per le immagini dei singoli reparti (senza accenti o spazi)
REPARTO_DIR_NAME = {
    'Edilizia e ferramenta': 'edilizia-ferramenta',
    'Elettricità': 'elettricita',
    'Elettrodomestici': 'elettrodomestici',
    'Elettroutensili': 'elettroutensili',
    'Idraulica': 'idraulica',
    'Utensili manuali': 'utensili-manuali',
    'Vernici e colori': 'vernici-colori',
    'Vario': 'vario'
}

# Mappatura Sottocategorie note per mantenere la coerenza dei codici (3 lettere)
SUBCAT_PREFIX_MAP = {
    'press control': 'PRC',
    'chiavi': 'CHI',
    'frigoriferi e congelatori': 'FRI',
    'prese e interruttori': 'PRE',
    'trapani e avvitatori': 'TRA',
    'pennelli e rulli': 'PEN',
    'articoli per la casa': 'CAS',
    'ferramenta varia': 'FER',
    'smerigliatrici': 'SME',
    'tubi e raccordi': 'TUB'
}

# Sottocategorie predefinite per i menù a tendine guidati
SOTTOCATEGORIE_REPARTI = {
    'Edilizia e ferramenta': [
        'Viti e bulloneria',
        'Tasselli e ancoraggi',
        'Cemento e malte',
        'Ferramenta varia',
        'Catene e corde',
        'Lucchetti e serrature',
        'Scarpe',
        'Altro'
    ],
    'Elettricità': [
        'Cavi e prolunghe',
        'Prese e interruttori',
        'Illuminazione',
        'Multiprese',
        'Fusibili e quadri',
        'Torce e pile',
        'Altro'
    ],
    'Elettrodomestici': [
        'Frigoriferi e congelatori',
        'Lavatrici e asciugatrici',
        'Lavastoviglie',
        'Cucine',
        'Televisori',
        'Piccoli elettrodomestici',
        'Elettrodomestici da incasso',
        'Climatizzazione',
        'Riscaldamento',
        'Ricambi e accessori',
        'Cura della persona',
        'Altro'
    ],
    'Elettroutensili': [
        'Trapani e avvitatori',
        'Seghe',
        'Levigatrici',
        'Smerigliatrici',
        'Batterie e caricabatterie',
        'Accessori',
        'Altro'
    ],
    'Idraulica': [
        'Tubi e raccordi',
        'Rubinetteria',
        'Sifoni e scarichi',
        'Guarnizioni',
        'Pompe',
        'Teflon e sigillanti',
        'Press control',
        'Accessori doccia',
        'Altro'
    ],
    'Utensili manuali': [
        'Chiavi',
        'Cacciaviti e bit',
        'Pinze e tronchesi',
        'Martelli e mazze',
        'Metri e livelle',
        'Taglio',
        'Set e valigette',
        'Altro'
    ],
    'Vernici e colori': [
        'Pitture murali',
        'Smalti e vernici legno',
        'Pennelli e rulli',
        'Solventi e diluenti',
        'Nastri e teli protettivi',
        'Stucchi e decorazioni',
        'Tinte decorative',
        'Vernici speciali',
        'Altro'
    ],
    'Vario': [
        'Articoli per la casa',
        'Contenitori e organizzazione',
        'Cancelleria e ufficio',
        'Pulizia',
        'Articoli stagionali',
        'Giardinaggio',
        'Altro'
    ]
}

def get_subcategory_code(subcat_name):
    if pd.isna(subcat_name) or str(subcat_name).strip() == '':
        return 'GEN'
    
    clean_name = str(subcat_name).strip().lower()
    
    # Cerca nella mappatura predefinita
    if clean_name in SUBCAT_PREFIX_MAP:
        return SUBCAT_PREFIX_MAP[clean_name]
    
    # Altrimenti, prendi le prime 3 lettere escludendo spazi e speciali
    sub_clean = re.sub(r'[^a-zA-Z]+', '', clean_name).upper()
    sub_code = sub_clean[:3] if len(sub_clean) >= 3 else (sub_clean + "XXX")[:3]
    return sub_code

def clean_price(price_val):
    if pd.isna(price_val) or str(price_val).strip() == '':
        return 0.0
    p_str = str(price_val).replace('€', '').replace('$', '').strip()
    p_str = p_str.replace(',', '.')
    match = re.search(r'\d+(?:\.\d+)?', p_str)
    if match:
        return float(match.group(0))
    return 0.0

def parse_variants(variants_val):
    if pd.isna(variants_val) or str(variants_val).strip() == '':
        return None
    
    val_str = str(variants_val).strip()
    if ':' not in val_str:
        return None
    
    try:
        parts = val_str.split(':', 1)
        label = parts[0].strip()
        options_part = parts[1].strip()
        
        options_list = []
        raw_options = options_part.split('|')
        for raw_opt in raw_options:
            if '=' in raw_opt:
                opt_parts = raw_opt.split('=', 1)
                opt_val = opt_parts[0].strip()
                opt_price = clean_price(opt_parts[1])
                options_list.append({
                    "value": opt_val,
                    "price": opt_price
                })
        
        if options_list:
            return {
                "label": label,
                "options": options_list
            }
    except Exception as e:
        print(f"Errore nel parsing della variante '{val_str}': {e}")
        
    return None

def format_variants_for_excel(variants_obj):
    if not variants_obj or not isinstance(variants_obj, dict):
        return ''
    
    label = variants_obj.get('label', '')
    options = variants_obj.get('options', [])
    if not label or not options:
        return ''
    
    opt_strings = []
    for opt in options:
        val = opt.get('value', '')
        price = opt.get('price', 0.0)
        try:
            price_float = float(price)
        except:
            price_float = clean_price(price)
        opt_strings.append(f"{val} = {price_float:.2f}")
        
    return f"{label}: " + " | ".join(opt_strings)

def get_simplified_image_value(image_path, code, reparto_name):
    if not image_path:
        return ""
    if image_path.startswith('data:'):
        return '[Immagine in Base64]'
    if image_path.startswith('http://') or image_path.startswith('https://'):
        return image_path
    
    # Sostituiamo backslash con forward slash per coerenza
    normalized_path = image_path.replace('\\\\', '/').replace('\\', '/')
    safe_reparto_dir = REPARTO_DIR_NAME.get(reparto_name, 'vario')
    standard_prefix = f"assets/img/{safe_reparto_dir}/"
    old_standard_prefix = f"assets/img/{reparto_name}/"
    
    # Se è nella nuova cartella web-safe
    if normalized_path.startswith(standard_prefix):
        filename = os.path.basename(normalized_path)
        base, ext = os.path.splitext(filename)
        # Se il nome corrisponde al codice prodotto o codice_indice
        if base == code or base.startswith(f"{code}_"):
            return base # es. 'ED-FRI-001' o 'ED-FRI-001_1'
        else:
            return filename # es. 'trapano.jpg'
    # Se era nella vecchia cartella intera
    elif normalized_path.startswith(old_standard_prefix):
        filename = os.path.basename(normalized_path)
        base, ext = os.path.splitext(filename)
        if base == code or base.startswith(f"{code}_"):
            return base
        else:
            return filename
    else:
        return image_path # Mantiene il percorso originale se è esterno

def get_simplified_images_string(prod, reparto_name):
    images = prod.get('images', [])
    if not images:
        single_img = prod.get('image', '')
        if single_img:
            images = [single_img]
            
    simplified_parts = []
    for img_path in images:
        simplified_val = get_simplified_image_value(img_path, prod.get('code', ''), reparto_name)
        if simplified_val:
            simplified_parts.append(simplified_val)
            
    return " | ".join(simplified_parts)

def compress_and_convert_to_webp(file_path, max_width=800, quality=80):
    """
    Ottimizza un'immagine: la converte in WebP, la ridimensiona se supera max_width 
    e riduce il peso del file per velocizzare il caricamento del sito web.
    """
    try:
        from PIL import Image
        if not os.path.exists(file_path):
            return file_path
        
        base, ext = os.path.splitext(file_path)
        ext_lower = ext.lower().lstrip('.')
        
        # Ignora file non grafici o formati vettoriali/animati
        if ext_lower in ['svg', 'gif']:
            return file_path
            
        with Image.open(file_path) as img:
            # Rileva e converte formati non RGB (es. CMYK)
            if img.mode in ('RGBA', 'LA') and ext_lower != 'png' and ext_lower != 'webp':
                pass
            elif img.mode != 'RGB' and img.mode != 'RGBA':
                img = img.convert('RGB')
                
            # Ridimensionamento proporzionale intelligente
            width, height = img.size
            if width > max_width:
                ratio = max_width / float(width)
                new_height = int(height * ratio)
                img = img.resize((max_width, new_height), Image.Resampling.LANCZOS)
                print(f"     [INFO] Ridimensionata immagine da {width}x{height}px a {max_width}x{new_height}px")
            
            webp_path = f"{base}.webp"
            img.save(webp_path, 'WEBP', quality=quality)
            print(f"     [OK] Immagine compressa salvata come WebP: '{webp_path}'")
            
            # Se il file originale non era un webp, rimuoviamolo per non sprecare spazio su GitHub
            if ext_lower != 'webp' and os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    print(f"     [INFO] Rimosso file originale non ottimizzato: '{file_path}'")
                except Exception as e:
                    print(f"     [Nota] Impossibile rimuovere il file originale: {e}")
                    
            return webp_path
    except Exception as e:
        print(f"     [ERRORE] Errore durante l'ottimizzazione dell'immagine: {e}")
        return file_path

def download_image_from_url(url, output_dir, code):
    # Crea la directory se non esiste
    os.makedirs(output_dir, exist_ok=True)
    
    # Determina l'estensione ipotetica dall'URL
    parsed_path = urllib.parse.urlparse(url).path
    _, url_ext = os.path.splitext(parsed_path)
    ext = url_ext.lstrip('.').lower()
    if ext not in ['jpg', 'jpeg', 'png', 'webp', 'gif']:
        ext = 'jpg' # Default
    if ext == 'jpeg':
        ext = 'jpg'
        
    filename = f"{code}.{ext}"
    file_path = os.path.join(output_dir, filename)
    reparto_dir_slug = os.path.basename(output_dir)
    rel_path = f"assets/img/{reparto_dir_slug}/{filename}"
    
    try:
        print(f"  -> Tentativo di download da: {url} ...")
        req = urllib.request.Request(
            url, 
            headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'}
        )
        with urllib.request.urlopen(req, timeout=10) as response:
            content_type = response.headers.get('content-type')
            if content_type:
                guessed_ext = mimetypes.guess_extension(content_type)
                if guessed_ext:
                    new_ext = guessed_ext.lstrip('.').lower()
                    if new_ext == 'jpeg':
                        new_ext = 'jpg'
                    if new_ext in ['jpg', 'png', 'webp', 'gif'] and new_ext != ext:
                        ext = new_ext
                        filename = f"{code}.{ext}"
                        file_path = os.path.join(output_dir, filename)
                        rel_path = f"assets/img/{reparto_dir_slug}/{filename}"
            
            with open(file_path, 'wb') as out_file:
                out_file.write(response.read())
            
            print(f"     [OK] Immagine scaricata e salvata in '{rel_path}'")
            
            # Ottimizza l'immagine scaricata convertendola in WebP
            optimized_path = compress_and_convert_to_webp(file_path)
            # Sostituiamo gli slash per sicurezza web
            optimized_rel_path = optimized_path.replace('\\', '/')
            if '/workspace/scratch/manutenzione/' in optimized_rel_path:
                optimized_rel_path = optimized_rel_path.replace('/workspace/scratch/manutenzione/', '')
            
            return optimized_rel_path
    except Exception as e:
        print(f"     [ERRORE] Impossibile scaricare l'immagine da Internet: {e}")
        if "Name or service not known" in str(e) or "Temporary failure in name resolution" in str(e) or "timed out" in str(e).lower() or "connection refused" in str(e).lower():
            print("     (Nota: Se ti trovi in un ambiente offline come il sandbox di test, questo errore è normale. Sul tuo PC Windows funzionerà correttamente!)")
        # In caso di errore di rete, ritorniamo comunque il percorso relativo ipotizzato
        # in formato .webp poiché verrà creata così una volta scaricata dal tuo PC
        hypothetical_rel_path = f"assets/img/{reparto_dir_slug}/{code}.webp"
        print(f"     [FALLBACK] Verrà comunque associato il percorso ottimizzato: '{hypothetical_rel_path}'")
        return hypothetical_rel_path

def save_json_with_backup(data, file_path):
    """
    Salva i dati in formato JSON creando prima un file di backup (.bak) di sicurezza.
    """
    try:
        if os.path.exists(file_path):
            backup_path = file_path + ".bak"
            shutil.copy2(file_path, backup_path)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Errore durante il salvataggio o backup del file {file_path}: {e}")
        return False

def import_excel_to_json(excel_file, data_dir):
    print(f"Lettura file Excel: {excel_file}...")
    try:
        xls = pd.ExcelFile(excel_file)
    except Exception as e:
        print(f"Errore nel caricamento del file Excel: {e}")
        return

    # Per ogni foglio (che rappresenta un reparto)
    for sheet_name in xls.sheet_names:
        if sheet_name not in REPARTO_MAP:
            if sheet_name != 'Sottocategorie':
                print(f"Ignorato foglio non standard: '{sheet_name}'")
            continue
            
        print(f"\nElaborazione reparto '{sheet_name}'...")
        df = pd.read_excel(xls, sheet_name=sheet_name)
        
        # Carichiamo il JSON esistente per preservare i Base64 o vecchi dati
        filename = REPARTO_MAP[sheet_name]
        json_path = os.path.join(data_dir, filename)
        existing_images = {}
        
        if os.path.exists(json_path):
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    old_data = json.load(f)
                    for p in old_data:
                        if p.get('name') and p.get('image'):
                            # Sotto-array images o singola stringa image
                            existing_images[p['name'].strip().lower()] = p.get('images', [p['image']])
            except Exception as e:
                print(f"Nota: Impossibile caricare il JSON esistente per mappare le immagini: {e}")
        
        # Rileviamo tutti i codici già definiti in questo Excel per evitare di duplicarli
        # e per determinare il contatore iniziale corretto per ogni sottocategoria
        assigned_codes = {}
        
        # Prima passata: leggiamo i codici esistenti
        for idx, row in df.iterrows():
            code_val = row.get('Codice Prodotto')
            if not pd.isna(code_val) and str(code_val).strip() != '':
                code_str = str(code_val).strip()
                match = re.match(r'^([A-Z]{2}-[A-Z]{3})-(\d{3})$', code_str)
                if match:
                    prefix = match.group(1)
                    num = int(match.group(2))
                    if prefix not in assigned_codes:
                        assigned_codes[prefix] = []
                    assigned_codes[prefix].append(num)
        
        products_list = []
        for idx, row in df.iterrows():
            name = str(row.get('Nome Prodotto', '')).strip()
            if not name or pd.isna(row.get('Nome Prodotto')):
                continue # Salta righe vuote
                
            # Se il prezzo non è ancora stato specificato, la riga è considerata
            # una bozza non pronta: la ignoriamo e non la scriviamo nel JSON
            # pubblicato sul sito. Resta comunque nel file Excel così puoi
            # completarla con calma (vedi anche export_json_to_excel).
            raw_price = row.get('Prezzo Base')
            if pd.isna(raw_price) or str(raw_price).strip() == '':
                print(f"  -> Prodotto '{name}' ignorato: prezzo non specificato (bozza non ancora pronta).")
                continue

            subcategory = str(row.get('Sottocategoria', '')).strip()
            price_base = clean_price(raw_price)
            description = str(row.get('Descrizione', '')).strip() if not pd.isna(row.get('Descrizione')) else ""
            
            # Determinazione del Codice Prodotto intelligente
            code_val = row.get('Codice Prodotto')
            rep_prefix = REPARTO_PREFIX.get(sheet_name, 'XX')
            sub_prefix = get_subcategory_code(subcategory)
            counter_key = f"{rep_prefix}-{sub_prefix}"
            
            if not pd.isna(code_val) and str(code_val).strip() != '':
                code = str(code_val).strip()
            else:
                # Generiamo automaticamente il codice progressivo libero!
                if counter_key not in assigned_codes:
                    assigned_codes[counter_key] = []
                
                next_num = 1
                if assigned_codes[counter_key]:
                    next_num = max(assigned_codes[counter_key]) + 1
                
                assigned_codes[counter_key].append(next_num)
                code = f"{counter_key}-{next_num:03d}"
                print(f"  -> Generato codice automatico per '{name}': {code}")
            
            # Gestione intelligente delle immagini multiple
            image_val = row.get('Immagine')
            image_paths = []
            
            if not pd.isna(image_val):
                image_val_str = str(image_val).strip()
                if image_val_str == '[Immagine in Base64]':
                    # Recupera l'originario dal file JSON esistente
                    lookup_key = name.lower()
                    if lookup_key in existing_images:
                        image_paths = existing_images[lookup_key]
                elif image_val_str and image_val_str != 'nan' and image_val_str != '':
                    # Splitta le immagini col pipe | per consentire immagini multiple
                    parts = [p.strip() for p in image_val_str.split('|') if p.strip()]
                    for idx_img, part in enumerate(parts):
                        img_code = code if idx_img == 0 else f"{code}_{idx_img}"
                        
                        # SE E' UN URL WEB! (http:// o https://)
                        if part.startswith('http://') or part.startswith('https://'):
                            reparto_dir_slug = REPARTO_DIR_NAME.get(sheet_name, 'vario')
                            output_dir = os.path.join('assets', 'img', reparto_dir_slug)
                            # Gestione sandbox
                            if not os.path.exists('assets') and os.path.exists('/workspace/scratch/manutenzione/assets'):
                                output_dir = os.path.join('/workspace/scratch/manutenzione', 'assets', 'img', reparto_dir_slug)
                            
                            downloaded_path = download_image_from_url(part, output_dir, img_code)
                            if downloaded_path:
                                image_paths.append(downloaded_path)
                                
                        # Se non contiene slash/backslash, è un nome file o codice semplificato!
                        elif '/' not in part and '\\\\' not in part:
                            reparto_dir_slug = REPARTO_DIR_NAME.get(sheet_name, 'vario')
                            if '.' in part:
                                # Ha già l'estensione (es. 'ED-FRI-001.webp')
                                image_paths.append(f"assets/img/{reparto_dir_slug}/{part}")
                            else:
                                # È solo il codice o nome base (es. 'ED-FRI-001')
                                found_path = None
                                local_reparto_dir = os.path.join('assets', 'img', reparto_dir_slug)
                                if not os.path.exists('assets') and os.path.exists('/workspace/scratch/manutenzione/assets'):
                                    local_reparto_dir = os.path.join('/workspace/scratch/manutenzione', 'assets', 'img', reparto_dir_slug)
                                if os.path.exists(local_reparto_dir):
                                    for ext in ['.webp', '.jpg', '.png', '.jpeg', '.WEBP', '.JPG', '.PNG']:
                                        test_file = f"{part}{ext}"
                                        if os.path.exists(os.path.join(local_reparto_dir, test_file)):
                                            found_path = f"assets/img/{reparto_dir_slug}/{test_file}"
                                            break
                                if found_path:
                                    image_paths.append(found_path)
                                else:
                                    # Default a .webp se non trovato (formato v11/v12)
                                    image_paths.append(f"assets/img/{reparto_dir_slug}/{part}.webp")
                        else:
                            # È un percorso intero, normalizziamo il nome della cartella reparto per sicurezza web-safe
                            normalized_path = part.replace('\\\\', '/').replace('\\', '/')
                            for old_rep, new_slug in REPARTO_DIR_NAME.items():
                                old_pattern = f"assets/img/{old_rep}/"
                                if normalized_path.startswith(old_pattern):
                                    normalized_path = normalized_path.replace(old_pattern, f"assets/img/{new_slug}/")
                                    break
                            image_paths.append(normalized_path)
            
            # Se la cella dell'immagine era vuota, proviamo ad auto-associarla
            # cercando code.webp, code_1.webp, code_2.webp etc. nella cartella corretta!
            if not image_paths:
                reparto_dir_slug = REPARTO_DIR_NAME.get(sheet_name, 'vario')
                local_reparto_dir = os.path.join('assets', 'img', reparto_dir_slug)
                if not os.path.exists('assets') and os.path.exists('/workspace/scratch/manutenzione/assets'):
                    local_reparto_dir = os.path.join('/workspace/scratch/manutenzione', 'assets', 'img', reparto_dir_slug)
                
                if os.path.exists(local_reparto_dir):
                    primary_path = None
                    for ext in ['.webp', '.jpg', '.png', '.jpeg', '.WEBP', '.JPG', '.PNG']:
                        test_file = f"{code}{ext}"
                        if os.path.exists(os.path.join(local_reparto_dir, test_file)):
                            primary_path = f"assets/img/{reparto_dir_slug}/{test_file}"
                            break
                    if primary_path:
                        image_paths.append(primary_path)
                        print(f"  -> Auto-associata immagine principale per '{name}': {primary_path}")
                        
                        # Cerca immagini secondarie sequenziali (es. code_1, code_2...)
                        idx_sec = 1
                        while True:
                            sec_found = False
                            for ext in ['.webp', '.jpg', '.png', '.jpeg', '.WEBP', '.JPG', '.PNG']:
                                test_file = f"{code}_{idx_sec}{ext}"
                                if os.path.exists(os.path.join(local_reparto_dir, test_file)):
                                    sec_path = f"assets/img/{reparto_dir_slug}/{test_file}"
                                    image_paths.append(sec_path)
                                    print(f"     -> Auto-associata immagine secondaria: {sec_path}")
                                    sec_found = True
                                    break
                            if not sec_found:
                                break
                            idx_sec += 1

            variants = parse_variants(row.get('Varianti'))
            
            # Definiamo immagine primaria ed array di immagini totali
            primary_image = image_paths[0] if image_paths else ""
            
            product_obj = {
                "code": code,
                "name": name,
                "category": subcategory,
                "price": price_base,
                "image": primary_image,
                "images": image_paths,
                "description": description
            }
            
            if variants:
                product_obj["variants"] = variants
                
            products_list.append(product_obj)
            
        # Salva in formato JSON con backup di sicurezza
        if save_json_with_backup(products_list, json_path):
            print(f"File JSON aggiornato con successo: '{json_path}' ({len(products_list)} prodotti)")

def load_existing_draft_rows(excel_file):
    """
    Rilegge il file Excel così com'è PRIMA di rigenerarlo, e recupera le righe
    "bozza": prodotti con un nome ma senza prezzo specificato. import_excel_to_json
    ignora sempre queste righe (non finiscono mai nei JSON pubblicati), quindi se
    export_json_to_excel si limitasse a ricostruire i fogli solo dai JSON, --export
    le farebbe sparire per sempre. Le recuperiamo qui per poterle riscrivere tali e
    quali nel nuovo file.
    """
    drafts = {name: [] for name in REPARTO_MAP}
    if not os.path.exists(excel_file):
        return drafts

    try:
        xls = pd.ExcelFile(excel_file)
    except Exception as e:
        print(f"Nota: impossibile rileggere l'Excel esistente per recuperare le bozze senza prezzo: {e}")
        return drafts

    for sheet_name in xls.sheet_names:
        if sheet_name not in REPARTO_MAP:
            continue
        try:
            df = pd.read_excel(xls, sheet_name=sheet_name)
        except Exception as e:
            print(f"Nota: impossibile leggere il foglio '{sheet_name}' per recuperare le bozze: {e}")
            continue

        for _, row in df.iterrows():
            name = str(row.get('Nome Prodotto', '')).strip()
            if not name or pd.isna(row.get('Nome Prodotto')):
                continue
            raw_price = row.get('Prezzo Base')
            if not (pd.isna(raw_price) or str(raw_price).strip() == ''):
                continue  # ha già un prezzo: verrà rigenerata regolarmente dal JSON

            def cell_str(val):
                return '' if pd.isna(val) else str(val).strip()

            drafts[sheet_name].append({
                'code': cell_str(row.get('Codice Prodotto')),
                'name': name,
                'category': cell_str(row.get('Sottocategoria')),
                'image': cell_str(row.get('Immagine')),
                'description': cell_str(row.get('Descrizione')),
                'variants': cell_str(row.get('Varianti')),
            })

    return drafts

def export_json_to_excel(excel_file, data_dir):
    print(f"Inizio esportazione dei file JSON da '{data_dir}' verso l'Excel '{excel_file}'...")
    draft_rows_by_reparto = load_existing_draft_rows(excel_file)
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    columns = [
        'Codice Prodotto',
        'Nome Prodotto',
        'Sottocategoria',
        'Prezzo Base',
        'Immagine',
        'Descrizione',
        'Varianti'
    ]
    
    # Stili
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    border_thin = Side(border_style='thin', color='D9D9D9')
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    # Copia delle sottocategorie predefinite per arricchirle dinamicamente
    sottocategorie_reali = copy.deepcopy(SOTTOCATEGORIE_REPARTI)
    
    # Per ogni reparto mappato, leggiamo i dati per raccogliere eventuali sottocategorie custom
    for reparto_name, json_filename in REPARTO_MAP.items():
        json_path = os.path.join(data_dir, json_filename)
        if os.path.exists(json_path):
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    products = json.load(f)
                    for prod in products:
                        subcat = prod.get('category', '').strip()
                        if subcat and subcat not in sottocategorie_reali[reparto_name]:
                            sottocategorie_reali[reparto_name].append(subcat)
            except Exception as e:
                print(f"Nota: Impossibile leggere {json_filename} per mappare le sottocategorie: {e}")
                
    # Ora generiamo i singoli fogli dei Reparti
    for reparto_name, json_filename in REPARTO_MAP.items():
        json_path = os.path.join(data_dir, json_filename)
        ws = wb.create_sheet(title=reparto_name)
        
        # Scrivi intestazione
        ws.append(columns)
        ws.row_dimensions[1].height = 25
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = cell_border
            
        added_names = set()

        if os.path.exists(json_path):
            print(f"Caricamento prodotti per il reparto '{reparto_name}' dal file '{json_filename}'...")
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    products = json.load(f)
            except Exception as e:
                print(f"Errore nel parsing del file {json_path}: {e}")
                products = []
                
            for prod in products:
                code = prod.get('code', '')
                name = prod.get('name', '')
                subcategory = prod.get('category', '')
                price = prod.get('price', 0.0)
                
                try:
                    price = float(price)
                except:
                    price = clean_price(price)
                    
                # Semplificazione intelligente della lista delle immagini
                img = get_simplified_images_string(prod, reparto_name)
                
                desc = prod.get('description', '')
                var_str = format_variants_for_excel(prod.get('variants'))
                
                row_data = [code, name, subcategory, price, img, desc, var_str]
                ws.append(row_data)
                if name:
                    added_names.add(name.strip().lower())
        else:
            print(f"Nessun file JSON trovato per il reparto '{reparto_name}' ({json_filename}). Creato foglio vuoto.")
            ws.append(['', 'Esempio Prodotto', 'Categoria Esempio', 10.00, '', '', ''])
            added_names.add('esempio prodotto')

        # Ripristina le bozze senza prezzo che erano già nell'Excel: import_excel_to_json
        # le ignora sempre, quindi non esistono nel JSON e andrebbero perse se non le
        # riscrivessimo qui. Se nel frattempo un prodotto è stato completato con un
        # prezzo (ed è quindi già presente tra i prodotti appena scritti), non lo
        # duplichiamo.
        for draft in draft_rows_by_reparto.get(reparto_name, []):
            if draft['name'].strip().lower() in added_names:
                continue
            row_data = [draft['code'], draft['name'], draft['category'], '', draft['image'], draft['description'], draft['variants']]
            ws.append(row_data)
            added_names.add(draft['name'].strip().lower())
            print(f"  -> Bozza senza prezzo mantenuta per '{reparto_name}': '{draft['name']}'")
            
        # Formatta celle dati
        for r_num in range(2, ws.max_row + 1):
            ws.row_dimensions[r_num].height = 20
            for col_num in range(1, len(columns) + 1):
                cell = ws.cell(row=r_num, column=col_num)
                cell.font = Font(name='Segoe UI', size=10)
                cell.border = cell_border
                
                if col_num in [1, 4]: # Codice, Prezzo
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                
                if col_num == 4:
                    cell.number_format = '[$€-2] #,##0.00'
                    
        # Ridimensiona colonne
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
            
        # Applica il menù a tendine per la Sottocategoria (Colonna C)
        headers_sub = list(sottocategorie_reali.keys())
        col_idx_sub = headers_sub.index(reparto_name) + 1
        col_letter_sub = openpyxl.utils.get_column_letter(col_idx_sub)
        num_items = len(sottocategorie_reali[reparto_name])
        formula_val = f"Sottocategorie!${col_letter_sub}$2:${col_letter_sub}${num_items + 1}"
        
        dv_sub = DataValidation(type='list', formula1=formula_val, allow_blank=True)
        dv_sub.error = "Seleziona una sottocategoria valida dalla lista"
        dv_sub.errorTitle = "Sottocategoria non valida"
        dv_sub.prompt = "Scegli una sottocategoria dalla lista"
        dv_sub.promptTitle = "Sottocategoria"
        
        ws.add_data_validation(dv_sub)
        dv_sub.add("C2:C200") # Valido fino a riga 200 per foglio
        
    # Scrittura del foglio Sottocategorie (Impostazioni per i menù a tendine)
    ws_sub = wb.create_sheet(title='Sottocategorie')
    headers_sub = list(sottocategorie_reali.keys())
    ws_sub.append(headers_sub)
    
    max_len = max(len(lst) for lst in sottocategorie_reali.values())
    for r in range(max_len):
        row_data = []
        for rep in headers_sub:
            lst = sottocategorie_reali[rep]
            if r < len(lst):
                row_data.append(lst[r])
            else:
                row_data.append(None)
        ws_sub.append(row_data)
        
    # Nascondiamo il foglio per non disturbare la visualizzazione dell'utente
    ws_sub.sheet_state = 'hidden'
            
    # Salva il file Excel
    try:
        wb.save(excel_file)
        print(f"\nSincronizzazione completata! Excel generato con successo: {excel_file}")
    except Exception as e:
        print(f"Errore nel salvataggio dell'Excel: {e}")

def run_global_optimization(data_dir, img_base_dir):
    """
    Scansiona tutte le cartelle di immagini reali, rinomina le directory dei reparti
    con nomi web-safe (es. 'Elettricità' -> 'elettricita'), converte tutte le immagini
    presenti in WebP compresso e aggiorna tutti i percorsi di riferimento nei file JSON.
    """
    print("\n==================================================")
    print("      AVVIO OTTIMIZZAZIONE GLOBALE DEL SITO")
    print("==================================================")
    
    if not os.path.exists(img_base_dir):
        print(f"[ERRORE] La cartella immagini '{img_base_dir}' non esiste.")
        return
        
    # Fase 1: Rinomina cartelle reparti con accenti/spazi in nomi Web-Safe
    print("\n[Fase 1] Controllo e migrazione nomi cartelle in formato Web-Safe...")
    for old_rep, clean_slug in REPARTO_DIR_NAME.items():
        old_path = os.path.join(img_base_dir, old_rep)
        new_path = os.path.join(img_base_dir, clean_slug)
        
        # Se esiste la vecchia cartella e non è ancora la nuova
        if os.path.exists(old_path) and old_rep != clean_slug:
            # Se esiste già la nuova cartella, unisci i file
            if os.path.exists(new_path):
                print(f"  -> Unione contenuti da '{old_path}' verso '{new_path}'...")
                for file_name in os.listdir(old_path):
                    shutil.move(os.path.join(old_path, file_name), os.path.join(new_path, file_name))
                os.rmdir(old_path)
            else:
                print(f"  -> Ridenominazione cartella: '{old_path}' -> '{new_path}'...")
                os.rename(old_path, new_path)
    
    # Fase 2: Converte tutte le immagini locali in WebP e le comprime
    print("\n[Fase 2] Compressione e conversione di tutte le immagini in WebP...")
    for reparto_slug in REPARTO_DIR_NAME.values():
        reparto_path = os.path.join(img_base_dir, reparto_slug)
        if os.path.exists(reparto_path):
            print(f"  Analisi immagini in '{reparto_path}'...")
            for filename in os.listdir(reparto_path):
                file_path = os.path.join(reparto_path, filename)
                if os.path.isfile(file_path):
                    _, ext = os.path.splitext(filename)
                    if ext.lower() in ['.jpg', '.jpeg', '.png']:
                        compress_and_convert_to_webp(file_path)
                        
    # Fase 3: Aggiorna tutti i riferimenti nei file JSON
    print("\n[Fase 3] Aggiornamento dei database JSON con i nuovi percorsi Web-Safe ed estensioni .webp...")
    for reparto_name, json_filename in REPARTO_MAP.items():
        json_path = os.path.join(data_dir, json_filename)
        if os.path.exists(json_path):
            print(f"  Aggiornamento '{json_filename}'...")
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    products = json.load(f)
                
                updated = False
                for prod in products:
                    images_field = prod.get('images', [])
                    single_image = prod.get('image', '')
                    
                    # Se non esiste images_field ma c'è single_image, creiamo l'array
                    if not images_field and single_image:
                        images_field = [single_image]
                        
                    new_images_field = []
                    for img_path in images_field:
                        if img_path and not img_path.startswith('data:') and not img_path.startswith('http'):
                            normalized = img_path.replace('\\\\', '/').replace('\\', '/')
                            
                            # 1. Sostituisce i vecchi nomi dei reparti con i nuovi slug web-safe
                            for old_rep, clean_slug in REPARTO_DIR_NAME.items():
                                old_pattern = f"assets/img/{old_rep}/"
                                if normalized.startswith(old_pattern):
                                    normalized = normalized.replace(old_pattern, f"assets/img/{clean_slug}/")
                                    break
                            
                            # 2. Sostituisce l'estensione dell'immagine con .webp se è stata convertita
                            base, ext = os.path.splitext(normalized)
                            if ext.lower() in ['.jpg', '.jpeg', '.png']:
                                rel_check_path = normalized.replace('assets/', '')
                                absolute_check_dir = os.path.join(img_base_dir, REPARTO_DIR_NAME.get(reparto_name, 'vario'))
                                webp_filename = f"{os.path.basename(base)}.webp"
                                webp_full_path = os.path.join(absolute_check_dir, webp_filename)
                                
                                if os.path.exists(webp_full_path) or not os.path.exists(img_base_dir):
                                    normalized = f"assets/img/{REPARTO_DIR_NAME.get(reparto_name, 'vario')}/{webp_filename}"
                            
                            new_images_field.append(normalized)
                        else:
                            new_images_field.append(img_path)
                    
                    if new_images_field != images_field or (single_image and not prod.get('image')):
                        prod['images'] = new_images_field
                        prod['image'] = new_images_field[0] if new_images_field else ""
                        updated = True
                
                if updated:
                    save_json_with_backup(products, json_path)
                    print(f"    [OK] File '{json_filename}' aggiornato e ottimizzato!")
            except Exception as e:
                print(f"    [ERRORE] Impossibile aggiornare il file {json_filename}: {e}")
                
    print("\n==================================================")
    print("  OTTIMIZZAZIONE COMPLETATA CON SUCCESSO! 🎉")
    print("  Ora tutte le immagini del tuo sito sono WebP")
    print("  e le cartelle dei reparti sono web-safe (no accenti/spazi).")
    print("==================================================")

def main():
    parser = argparse.ArgumentParser(description="Script di Sincronizzazione ed Ottimizzazione Catalogo Excel <-> JSON v12")
    parser.add_argument('--export', action='store_true', help="Esporta i file JSON correnti per popolare l'Excel")
    parser.add_argument('--optimize-all', action='store_true', help="Ottimizza tutte le cartelle immagini e aggiorna i database a WebP")
    args = parser.parse_args()
    
    excel_file = 'catalogo_template.xlsx'
    
    # Se siamo nel sandbox, usiamo i percorsi corretti di test
    if not os.path.exists(excel_file) and os.path.exists('/workspace/scratch/catalogo_template-v12.xlsx'):
        excel_file = '/workspace/scratch/catalogo_template-v12.xlsx'
    elif not os.path.exists(excel_file):
        excel_file = '/workspace/scratch/catalogo_template-v12.xlsx'
        
    data_dir = 'assets/data'
    img_base_dir = os.path.join('assets', 'img')
    
    if not os.path.exists(data_dir):
        data_dir = '/workspace/scratch/manutenzione/assets/data'
        img_base_dir = '/workspace/scratch/manutenzione/assets/img'
        os.makedirs(data_dir, exist_ok=True)
        os.makedirs(img_base_dir, exist_ok=True)
        
    if args.optimize_all:
        run_global_optimization(data_dir, img_base_dir)
        # Dopo l'ottimizzazione globale, rigeneriamo l'Excel in automatico per mantenerlo allineato!
        export_json_to_excel(excel_file, data_dir)
    elif args.export:
        export_json_to_excel(excel_file, data_dir)
    else:
        import_excel_to_json(excel_file, data_dir)

if __name__ == '__main__':
    main()
